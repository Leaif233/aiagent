import json
import uuid
from datetime import datetime

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends

from db.sqlite_db import get_connection
from db.chroma_client import get_tickets_collection
from db.versioning import create_version_snapshot, log_audit
from core.auth import get_current_user, require_admin, CurrentUser
from db.fingerprint_store import delete_fingerprints_by_entity
from tasks.ticket_tasks import clean_ticket_task, approve_ticket_task
from models.schemas import BatchRequest, TicketUpdateRequest

router = APIRouter(prefix="/api/tickets", tags=["tickets"])


@router.post("/import")
async def import_tickets(file: UploadFile = File(...), user: CurrentUser = Depends(require_admin)):
    """Batch import tickets from Excel or JSON file."""
    content = await file.read()
    filename = file.filename or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower()

    raw_tickets = []

    if ext == "json":
        data = json.loads(content)
        raw_tickets = data if isinstance(data, list) else [data]
    elif ext in ("xlsx", "xls"):
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = dict(zip(headers, row))
            raw_tickets.append(item)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use .json or .xlsx")

    results = []
    conn = get_connection()
    try:
        for item in raw_tickets:
            ticket_id = uuid.uuid4().hex
            now = datetime.utcnow().isoformat()
            raw_content = item.get("raw_content") or json.dumps(item, ensure_ascii=False)
            ticket_number = item.get("ticket_number", f"TKT-{ticket_id[:6].upper()}")

            # Preserve pre-extracted fields if present
            phenomenon = item.get("phenomenon") or ""
            cause = item.get("cause") or ""
            solution = item.get("solution") or ""
            has_extracted = bool(phenomenon.strip() or cause.strip() or solution.strip())

            status = "待审核" if has_extracted else "待处理"

            conn.execute(
                """INSERT INTO tickets
                   (id, ticket_number, status, raw_content,
                    phenomenon, cause, solution, created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (ticket_id, ticket_number, status, raw_content,
                 phenomenon, cause, solution, now, now),
            )

            task_id = None
            if not has_extracted:
                task = clean_ticket_task.delay(ticket_id)
                task_id = task.id

            results.append({
                "ticket_id": ticket_id,
                "ticket_number": ticket_number,
                "task_id": task_id,
                "status": status,
            })

        conn.commit()
    finally:
        conn.close()

    return {"imported": len(results), "items": results}


@router.get("")
async def list_tickets(status: str = None, search: str = None, page: int = 1, page_size: int = 20, user: CurrentUser = Depends(get_current_user)):
    """List tickets with optional status filter, search, and pagination."""
    conn = get_connection()
    try:
        where_clauses = []
        params = []

        if status:
            where_clauses.append("status = ?")
            params.append(status)
        if search:
            where_clauses.append(
                "(ticket_number LIKE ? OR phenomenon LIKE ? OR cause LIKE ? OR solution LIKE ?)"
            )
            params.extend([f"%{search}%"] * 4)

        where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

        total = conn.execute(
            f"SELECT COUNT(*) as c FROM tickets{where_sql}", params
        ).fetchone()["c"]

        offset = (page - 1) * page_size
        rows = conn.execute(
            f"SELECT id, ticket_number, status, phenomenon, created_at, updated_at "
            f"FROM tickets{where_sql} ORDER BY updated_at DESC LIMIT ? OFFSET ?",
            params + [page_size, offset],
        ).fetchall()

        return {"items": [dict(r) for r in rows], "total": total, "page": page, "page_size": page_size}
    finally:
        conn.close()


@router.post("/batch")
async def batch_tickets(body: BatchRequest, user: CurrentUser = Depends(require_admin)):
    """Batch approve/reject/delete tickets."""
    if body.action not in ("approve", "reject", "delete") or not body.ids:
        raise HTTPException(status_code=400, detail="Invalid action or empty ids")

    success = 0
    failed = 0
    for ticket_id in body.ids:
        try:
            if body.action == "approve":
                await approve_ticket(ticket_id, user)
            elif body.action == "reject":
                await reject_ticket(ticket_id, user)
            elif body.action == "delete":
                await delete_ticket(ticket_id, user)
            success += 1
        except Exception:
            failed += 1

    return {"success": success, "failed": failed}


@router.get("/{ticket_id}")
async def get_ticket(ticket_id: str, user: CurrentUser = Depends(get_current_user)):
    """Get single ticket detail including raw and extracted fields."""
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT * FROM tickets WHERE id = ?", (ticket_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Ticket not found")
        return dict(row)
    finally:
        conn.close()


@router.patch("/{ticket_id}")
async def update_ticket(ticket_id: str, body: TicketUpdateRequest, user: CurrentUser = Depends(require_admin)):
    """Update extracted fields (editor saves)."""
    create_version_snapshot("ticket", ticket_id, user.username, "update")
    log_audit("ticket", ticket_id, "update", user.username)
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT id FROM tickets WHERE id = ?", (ticket_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Ticket not found")

        now = datetime.utcnow().isoformat()
        updates = body.model_dump(exclude_none=True)
        if not updates:
            raise HTTPException(status_code=400, detail="No valid fields to update")

        sets = [f"{k} = ?" for k in updates]
        vals = list(updates.values())
        sets.append("updated_at = ?")
        vals.append(now)
        vals.append(ticket_id)

        conn.execute(
            f"UPDATE tickets SET {', '.join(sets)} WHERE id = ?", vals
        )
        conn.commit()
        return {"status": "updated", "ticket_id": ticket_id}
    finally:
        conn.close()


@router.patch("/{ticket_id}/approve")
async def approve_ticket(ticket_id: str, user: CurrentUser = Depends(require_admin)):
    """Approve ticket: dispatch async task for embedding + indexing."""
    create_version_snapshot("ticket", ticket_id, user.username, "approve")
    log_audit("ticket", ticket_id, "approve", user.username)
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT id, phenomenon FROM tickets WHERE id = ?", (ticket_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Ticket not found")

        if not (row["phenomenon"] or "").strip():
            raise HTTPException(status_code=400, detail="No extracted content to embed")

        now = datetime.utcnow().isoformat()
        conn.execute(
            "UPDATE tickets SET status = '审核通过处理中', updated_at = ? WHERE id = ?",
            (now, ticket_id),
        )
        conn.commit()

        task = approve_ticket_task.delay(ticket_id)
        return {"status": "approved", "ticket_id": ticket_id, "task_id": task.id}
    finally:
        conn.close()


@router.patch("/{ticket_id}/reject")
async def reject_ticket(ticket_id: str, user: CurrentUser = Depends(require_admin)):
    """Reject ticket: set status to 已驳回."""
    create_version_snapshot("ticket", ticket_id, user.username, "reject")
    log_audit("ticket", ticket_id, "reject", user.username)
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT id FROM tickets WHERE id = ?", (ticket_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Ticket not found")

        now = datetime.utcnow().isoformat()
        conn.execute(
            "UPDATE tickets SET status = '已驳回', updated_at = ? WHERE id = ?",
            (now, ticket_id),
        )
        conn.commit()

        # Remove from ChromaDB
        collection = get_tickets_collection()
        try:
            collection.delete(ids=[ticket_id])
        except Exception:
            pass

        delete_fingerprints_by_entity("ticket", ticket_id)
        return {"status": "rejected", "ticket_id": ticket_id}
    finally:
        conn.close()


@router.post("/{ticket_id}/reclean")
async def reclean_ticket(ticket_id: str, user: CurrentUser = Depends(require_admin)):
    """Re-trigger LLM extraction for a ticket."""
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT id FROM tickets WHERE id = ?", (ticket_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Ticket not found")

        now = datetime.utcnow().isoformat()
        conn.execute(
            "UPDATE tickets SET status = '待处理', updated_at = ? WHERE id = ?",
            (now, ticket_id),
        )
        conn.commit()

        task = clean_ticket_task.delay(ticket_id)
        return {"status": "recleaning", "ticket_id": ticket_id, "task_id": task.id}
    finally:
        conn.close()


@router.delete("/{ticket_id}")
async def delete_ticket(ticket_id: str, user: CurrentUser = Depends(require_admin)):
    """Delete ticket and remove from ChromaDB."""
    create_version_snapshot("ticket", ticket_id, user.username, "delete")
    log_audit("ticket", ticket_id, "delete", user.username)
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT id FROM tickets WHERE id = ?", (ticket_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Ticket not found")

        # Remove from ChromaDB
        collection = get_tickets_collection()
        try:
            collection.delete(ids=[ticket_id])
        except Exception:
            pass

        conn.execute("DELETE FROM tickets WHERE id = ?", (ticket_id,))
        conn.commit()
        delete_fingerprints_by_entity("ticket", ticket_id)
        return {"status": "deleted", "ticket_id": ticket_id}
    finally:
        conn.close()
